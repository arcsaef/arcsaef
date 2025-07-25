import json
import pandas
import re
import requests
import yaml
import unidecode
import numpy as np
from pyzotero    import zotero as pyzt
from docxtpl     import DocxTemplate
from openpyxl    import load_workbook
from datetime    import datetime
from collections import defaultdict, Counter
from itertools   import chain
from openpyxl.styles import Border, Side

# ensure that all details are printed
pandas.set_option('display.max_colwidth', None) 

# get a bibliography for a zotero Id
''' Get a single bibliographic record based on a Zotero id '''
def get_biblio(df, id_zotero):
    bib = df[df.ID_Zotero == id_zotero].Combined.to_string(index=False).strip()
    if bib == 'Series([], )':
        return 'None'
    else:
        return bib

# get an organisational bibliography and org bib count
def get_org_biblio(df, res_outputs, item_type):
    bib = df[df.ID_Zotero.isin(res_outputs[(res_outputs.itemType == 
                                            item_type)].key)].Combined
    if len(bib) > 0:
        return '\n\n'.join(sorted(bib)), len(bib)
    else:
        if item_type == 'plenary':
            return 'None reported', 0
        elif item_type == 'book':
            return 'Book: None reported', 0
        elif item_type == 'bookSection':
            return 'Chapter: None reported', 0
        else:
            return 'None reported', 0

# For a dictionary value list return entries a separate lines
'''  The main use for this is to construct a bibliography '''
def value_exists(d, key=None):
    if key is None:
        if d == ['Series([], )']:
            return None
        elif len(d) > 0:
            return '\n\n'.join(sorted(d)).replace("\n\nNone", "")
        else:
            return None
    else:
        item = d.get(key)
        if bool(item):
            return '\n\n'.join(sorted(item)).replace("\n\nNone", "")
        else:
            return None

''' Count total presentations less reports that have been tagged  
    DEPRECATED: Plenary is now entered as a type not as a tag '''
def unique_presentation(context_idv, presentation):
    tags = ['Public', 'Women', 'Industry', 'Ngo', 
            'Ats', 'Govt', 'Museum', 'Plenary']
    for tag in tags:
        if context_idv[tag] is not None:
            # Exclude tags from Presentation output type
            try:
                for o in context_idv[tag].split("\n\n"):
                    presentation = re.sub(o, '', presentation).lstrip('\n')
            except Exception as e:
                print(f"Presentation output type: {e}")
    return presentation

''' Count total reports less reports that have been tagged '''
def unique_report(context_idv, report):
    tags = ['Public', 'Women', 'Industry', 'Ngo', 
            'Ats', 'Govt', 'Museum', 'Plenary']
    for tag in tags:
        if context_idv[tag] is not None:
            # Exclude tags from report output type
            try:
                for o in context_idv[tag].split("\n\n"):
                    report = re.sub(o, '', report).lstrip('\n')
            except Exception as e:
                    print(f"Report output type: {e}")
    return report

''' Count a publication list '''    
def value_count(outputs):
    if outputs is None:
        return 0
    elif outputs == 'None':
        return 0
    elif len(outputs) == 0:
        return 0
    else:
        return len(str.split(outputs, '\n\n'))

''' Count SRI researcher type '''
def sri_researcher_count(org_list, sri_position):
    total = 0
    for x in org_list:# org_summary
        if sri_position in x:
            total += 1
    return total
    
# split the array of responses from Postman into distinct response objects
''' Split the Postman response into separate JSON responses '''
def split_response(bulk_response):
    responses = bulk_response[0].split('{"response":')
    responses_json = {}
    for i in range(1, len(responses)):
        # remove
        s1 = json.loads(responses[i].rstrip().rstrip(",'").rstrip(']').replace(',"messages":[{"code":"0","message":"OK"}]}', ''))
        responses_json[s1['dataInfo']['layout']] = s1

    return responses_json

# https://pyzotero.readthedocs.io/en/latest/#zotero.Zotero.collections
''' Returns a researcher-publication dataframe '''
def matched_library(library, ppl_hash):
    result      = pandas.DataFrame(columns=['key', 'itemType', 'title', 
                                   'rights', 'pubyr', 'name', 'id_person', 
                                   'tags', 'publicationTitle', 'project', 'logistics'])
    bare_result = pandas.DataFrame(columns=['key', 'itemType', 'title', 
                                   'rights', 'pubyr', 'publicationTitle'])
                          
    for x in library:
        item  = x['data']
        if 'date' in item:
            # Match 4 consecutive numbers in a string i.e. YYYY
            pubyr = ''.join(re.findall( '\d{4}', item['date']))  
        else:
            if item['itemType'] == 'podcast':
                pubyr = ''.join(re.findall( '\d{4}', item['accessDate']))
            else:
                pubyr = '2000-01-01'
        extra  = item.get('extra',  'saef: Missing extra')
        title  = item.get('title',  'saef: Missing title')
        rights = item.get('rights', 'saef: Missing rights')

        if 'publicationTitle' in x['data']:
            # item = x['data']
            pub_title = item['publicationTitle']
        else:
            pub_title = None

        # Item type shoehorn
        if item.get('presentationType', 'none').lower() == 'plenary':     
            item_type = 'plenary'
        elif item.get('itemType', 'none').lower() == 'map' or \
             item.get('itemType', 'none').lower() == 'software':
            item_type = 'ntro'
        elif item.get('itemType', 'none').lower() == 'blogPost' or \
             item.get('itemType', 'none').lower() == 'webpage'  or \
             item.get('itemType', 'none').lower() == 'magazineArticle':
            item_type = 'newspaperArticle'
        elif item.get('itemType', 'none').lower() == 'videoRecording':     
            item_type = 'film'
        elif item.get('itemType', 'none').lower() == 'podcast':     
            item_type = 'radioBroadcast'
        else:
            item_type = item['itemType']

        bare_result.loc[len(bare_result)] = [item['key'], item_type, title, rights, pubyr, pub_title]
        saef_author_list    = re.findall('saef:.*', extra)                # Use saef keyword as an identifier
        saef_project_list   = re.findall('project:.*', extra.lower())     # Use project keyword as an identifier
        saef_logistics_list = re.findall('logistics:.*', extra.lower())   # Use logistics keyword as an identifier
        proj_str = ''.join(saef_project_list).replace('project:', '').replace('Project:', '').replace('|', '').strip()
        logi_str = ''.join(saef_logistics_list).replace('logistics:', '').replace('Logistics:', '').replace('|', '').strip()

        for names in saef_author_list:
            names_tidy  = re.sub('saef:', '', names, flags=re.IGNORECASE)
            names_split = names_tidy.strip().split(';') # Create a list of saef names
            
            for name in names_split:
                # Find FileMaker person key
                id_person = ppl_hash.get(name.strip())
                # 'key', ['Item Type', 'Title', 'Journal', 'publication', 'rights', 'date', author, id]
                store = [item['key'], item_type, title, rights, pubyr, 
                        name.strip(), id_person, item['tags'], 
                        pub_title, proj_str.upper(), logi_str.upper()]
                result.loc[len(result)] = store
    # bare_result is a dataset that does not hold any author information.
    # it's intended use is for summary statistic queries.
    # result is a fully described library dataset the fully includes author participation.
    bare_result = bare_result[(bare_result.itemType != 'note') & (bare_result.itemType != 'attachment')]
    return [bare_result, result]

''' Define a SAEF person. Add additional person attributes here as needed '''
def person_construct(responses_json, rpt_yr):
    people   = {}
    ppl_hash = {}

    # iterate over responses_json['People_Detail']['data'] to build the base People data model
    for prsn in responses_json['People_Detail']['data']:
        # Find all projects and fte per per person in format, e.g. T1_001 (60)
        prsn_projs = []
        for prsn_proj in responses_json['people_Projects']['data']:
            if prsn_proj['fieldData']['IDf_Person'] == prsn['fieldData']['ID_Person']:
                # exclue on hold projects. how?
                if prsn_proj['fieldData']['Projects::State'] == 'Active':
                    prsn_projs.append(f"{prsn_proj['fieldData']['Projects::ProjectCode']} ({prsn_proj['fieldData']['FTE']})")
        # find all workshops per person
        prsn_wrkshps = []
        for prsn_wrkshp in responses_json['people_Workshops']['data']:
            if prsn_wrkshp['fieldData']['IDf_Person'] == prsn['fieldData']['ID_Person']:
                prsn_wrkshps.append(prsn_wrkshp['fieldData']['IDf_Workshop'])
        # find crossnode supervision. Using portal data, interesting
        crossnode = 'No'
        if prsn['portalData']['people_Supervision'] != []:
            for x in prsn['portalData']['people_Supervision']:
                    if x['people_Supervision::Crossnode'] == 'Yes':
                        crossnode = x['people_Supervision::Crossnode']
        # find supervisees
        supervises = []
        for s in responses_json['people_Supervision']['data']:
            if prsn['fieldData']['ID_Person'] == s['fieldData']['ID_Manager']:
                supervises.append(s['fieldData']['ID_Reportee'])

        # build person advisory/scar data
        prsn_advisory = [d for d in prsn['portalData']['people_Advisory']
                            if d['people_Advisory::YearTo'] == '' or 
                            int(d['people_Advisory::YearTo']) >= rpt_yr ] 

        for d in prsn_advisory:
            d['advisory_named'] = f"{prsn['fieldData']['LastName']}:  {d['people_Advisory::AdvisoryRole']}"    

        # build person data point
        people[prsn['fieldData']['ID_Person']] = {
            'Title':        prsn['fieldData']['Title'], 
            'FirstName':    prsn['fieldData']['FirstName'], 
            'LastName':     prsn['fieldData']['LastName'],
            'CareerStage':  prsn['fieldData']['CareerStage'], 
            'State':        prsn['fieldData']['State'], 
            'FTE':          prsn['fieldData']['FTE'],
            'Position':     prsn['fieldData']['Position'],  
            'Gender':       prsn['fieldData']['Gender'],
            'StartDate':    prsn['fieldData']['StartDate'],
            'EndDate':      prsn['fieldData']['EndDate'], 
            'Org':          prsn['fieldData']['Organisations 2::ShortName'], 
            'Organisation': prsn['fieldData']['Organisations 2::LongName'],  
            'Profile':      prsn['fieldData']['Profile'],
            'SAEFFunded':   prsn['fieldData']['SAEFFunded'], 
            'Consent':      prsn['fieldData']['Consent'],
            'Orcid':        prsn['fieldData']['ORCID'],
            'Email':        prsn['fieldData']['Email'],
            'Postnominal':  prsn['fieldData']['PostNominals'],
            'Role':         prsn['fieldData']['Role'],
            'Grants':       prsn['portalData']['people_Grants'], 
            'Training':     prsn['portalData']['people_Training'], 
            'Prizes':       prsn['portalData']['Prizes'],
            'StudentProjectTitle':  prsn['fieldData']['StudentProjectTitle'], 
            'Projects':             prsn_projs, 
            'Workshops':            prsn_wrkshps,
            'Supervises':           supervises, 
            'Advisory':             prsn_advisory, 
            'CrossnodeSupervision': crossnode,
            'Salutation': f"{prsn['fieldData']['Title']} {prsn['fieldData']['FirstName']} {prsn['fieldData']['LastName']} {prsn['fieldData']['PostNominals']}".strip() }
            # 'Scopusid': 0 if prsn['fieldData']['ORCID'] not in scopus['orcid'].to_list() else
            #                                     scopus[scopus.orcid == '0000-0001-7804-6648']['scopusid'].to_string(index=False) }

        ppl_hash[f"{str.replace(prsn['fieldData']['FirstName'], ' ', '')}{str.replace(prsn['fieldData']['LastName'], ' ', '')}"] = prsn['fieldData']['ID_Person']
        
    return people, ppl_hash

''' Basic data of SAEF project info held in FM '''        
def project_construct(responses_json, people):
    op = []
    op_bio = []

    for x in responses_json['people_Projects']['data']:
        if x['fieldData']['Role'] == 'Contact':
            px = people.get(x['fieldData']['IDf_Person'])
            if px != None:
                op_bio.append([x['fieldData']['Projects::ProjectCode'], f"{px['Title']} {px['FirstName']} {px['LastName']}", "-"])
    # UOW request matches ver.4 of 2024 organisation template     
        if x['fieldData']['Role'] == 'Manager':
            px = people.get(x['fieldData']['IDf_Person'])
            if px != None:
                for i, y in enumerate(op_bio):
                    if op_bio[i][0] == x['fieldData']['Projects::ProjectCode']:
                        # update the last column in th list with manager details where available
                        op_bio[i][2] = f"{px['Title']} {px['FirstName']} {px['LastName']}"       
     
    for proj in responses_json['Projects_Detail']['data']:
        if len(proj['fieldData']['ProjectCode']) > 0:
            op.append([proj['fieldData']['ProjectCode'], proj['fieldData']['ProjectAlias'], \
                proj['fieldData']['ProjectTitle'], proj['fieldData']['ProjectLeadOrganisation'],
                proj['fieldData']['Status']])
    op_df = pandas.DataFrame(op_bio, columns = ['ProjectCode', 'Contact', 'Manager']).set_index(['ProjectCode']).sort_index()
    op_bio_df = pandas.DataFrame(op, columns = ['ProjectCode', 'ProjectAlias', 'ProjectTitle', \
                                        'ProjectLeadOrganisation', 'Status']).set_index(['ProjectCode']).sort_index()
    projects = op_df.merge(op_bio_df, on='ProjectCode')
    return projects


''' Retrieve a SAEF library from Zotero in 1 of 3 ways
    1. Use a data file, get_saef_library(json_library='some_json_file.json')
    2. Use the API to retrive the entire SAEF library, get_saef_library()
    3. Use the API to retrieve a year's worth of data, get_saef_library(yr='2023) 
    '''
def get_saef_library(yr=1, config_file='config/reporting.yaml'):
    with open(config_file, 'r') as file:
        cf = yaml.safe_load(file)

    zot = pyzt.Zotero(library_id = cf['library']['id'], \
                      library_type=cf['library']['type'], api_key=cf['library']['key'])

    if yr > 1:
        # use api to retrieve a years worth at top-level
        # Flaky, retrieves n=100 max! ToDo: Fix limit bug
        zot.add_parameters(qmode="titleCreatorYear")
        saef_library = zot.top(q=yr)
    if yr == 0:
        # use api to retreive all (i.e. the entire library) at top-level
        # it takes ~25 secs. use sparingly
        saef_library = zot.everything(zot.top())
    else:
        # use a file to load a subset or the entire library
        with open(cf['data']['saef_library_json'], mode='r', encoding='utf8') as f:
            data = f.read()
            saef_library = json.loads(data)

    return saef_library

''' This functions seeks to extend the ContactList search in FM by provinfg a 
    comma seperated list of projects per researcher'''
def contact_list_projects(ppl_saef):
    contact_list = []
    for prsn in ppl_saef:
        if ppl_saef.get(prsn)['State'] and str.startswith(ppl_saef.get(prsn)['Consent'], 'Yes'):
            plist = re.sub(r'\(\d+\)| \(\)', '',','.join(sorted(ppl_saef.get(prsn)['Projects']))).rstrip(" '")                                           
            contact_list.append( f" {ppl_saef.get(prsn)['FirstName']}; {ppl_saef.get(prsn)['LastName']}; {plist}" )
    # ToDo: ideally contact_list should be sorted by LastName
    return contact_list

''' Define all reporting args '''
def get_rpt_args(config_file='config/reporting.yaml'):
    with open(config_file, 'r') as file:
        cf = yaml.safe_load(file)

    rpt_yr  = cf['report']['reporting_year']
    data    = cf['report']['organisations']
    organisations = json.loads(data) # convert string dictionary into a python dicationary

    return rpt_yr, organisations


''' Return a student thesis is reseracheris a student, blank otherwise'''
def get_thesis(ppl_saef, prsn):
    if ppl_saef[prsn]['CareerStage']  == 'Student' and len(ppl_saef[prsn]['StudentProjectTitle']) == 0:
        return 'None'
    else:
        return ppl_saef[prsn]['StudentProjectTitle']

''' '''
def has_thesis(ppl_saef, prsn):
    if ppl_saef[prsn]['CareerStage'] != 'Student':
        return 'n/a'
    else:
        if ppl_saef[prsn]['CareerStage']  == 'Student' and len(ppl_saef[prsn]['StudentProjectTitle']) == 0:
            return 'No'
        else:
            return 'Yes'
        
''' Define all the templated needed for KPI reporting'''
def load_templates(config_file='config/reporting.yaml'):
     # load templates
    with open(config_file, 'r') as file:
        cf = yaml.safe_load(file)

    doc_idv   = DocxTemplate(cf['templates']['idv_template_word'])
    doc_org   = DocxTemplate(cf['templates']['org_template_word'])
    templates = {'doc_idv': doc_idv, 'doc_org': doc_org}
    return templates

''' Load all data need to run KPI reporting defined defined in the YAML configuration file '''
def load_data(config_file='config/reporting.yaml'):      
    with open(config_file, 'r') as file:
        cf = yaml.safe_load(file)
    # import total json file created from Postman Filemaker collection
    with open(cf['data']['postman_responses'], mode='r', encoding='utf8') as f:
        bulk_response = f.readlines()
    # Load a zotero biliography 
    biblio = pandas.read_excel(cf['data']['saef_library_biblio'], sheet_name=str(cf['report']['reporting_year']))

    # Load scopus/orcid/name triples
    orcid = pandas.read_csv(cf['data']['orcid'], names=['orcid', 'name'])

    # Load data warehouse path
    wh = cf['data']['db']

    return bulk_response, biblio, orcid, wh

''' Returns  information if the individual is active and does not fall into
    any of these positions - Advisory, Intern, Ombudspeople, Program Staff, Visitor, Volunteer]'''
# KPI - Individual
def get_context_idv(org, people, id_prsn, res_outputs, bibliography, yr):
    global context_idv # global is required because get_context_idv is used by other functions in this file.
    prsn         = people.get(id_prsn)
    context_idv  = {}

    prsn_end_dt_chk = prsn['EndDate'].split("-")[0]
    if ( prsn_end_dt_chk == '' or int(prsn_end_dt_chk) >= int(yr) ) and \
        prsn['Position'] not in ['Advisory','Intern', 'Ombudspeople', 
                                 'Visitor', 'Volunteer'] and \
        prsn['Org'] == org:
        prsn_output                 = {}
        prsn_prize, scar, advisory  = [], [], []
        ngo, industry, govt, women  = [], [], [], []
        public, ats, museum, pro    = [], [], [], []
        superMasters, superAssoc    = [], []
        scar_named, advisory_named, prize_named = [], [], []
        superPostdoc, superPhd, superHons       = [], [], []

        context_idv['Salutation']            = prsn.get('Salutation')
        context_idv['Position']              = prsn.get('Position')
        context_idv['Organisation']          = prsn.get('Organisation')
        context_idv['StartDateDMY']          = datetime.strptime(prsn['StartDate'], '%Y-%m-%d').date().strftime('%d/%m/%Y') if len(prsn['StartDate']) > 1 else None
        context_idv['EndDateDMY']            = datetime.strptime(prsn['EndDate'],   '%Y-%m-%d').date().strftime('%d/%m/%Y') if len(prsn['EndDate'])   > 1 else None
        context_idv['Fte']                   = prsn.get('FTE')
        context_idv['Profile']               = prsn.get('Profile')
        context_idv['StudentProjectTitle']   = prsn.get('StudentProjectTitle')
        context_idv['Saef_funded']           = prsn.get('SAEFFunded')
        context_idv['ProjectCodeFTEList']    = ', '.join(prsn.get('Projects')) if any(prsn.get('Projects')) else None
        context_idv['CareerStage']           = prsn.get('CareerStage')
        context_idv['CrossnodeSupervision']  = prsn.get('CrossnodeSupervision')
        context_idv['Workshop']              = prsn.get('Workshop')
        context_idv['Supervision']           = prsn.get('Supervisee')
        context_idv['lastname']              = prsn.get('LastName') # needed for alpha sort by surname
        context_idv['ProfileURL']            = profile_exists(f"{prsn.get('FirstName')}-{prsn.get('LastName')}")

        if prsn.get('Prizes'):
            for prize in prsn.get('Prizes'):
                if prize['Prizes::Year'] == int(yr):
                    prsn_prize.append(prize['Prizes::Title'])
                    prize_named.append(f"{prsn.get('LastName')}: {prize['Prizes::Title']} [{yr}]")

        if prsn.get('Advisory'):
            for role in prsn.get('Advisory'):
                if "".__eq__(role['people_Advisory::YearTo']) or role['people_Advisory::YearTo'] >= int(yr):
                    role_yr      = f"{role['people_Advisory::AdvisoryRole']} [{role['people_Advisory::YearFrom']} - {role['people_Advisory::YearTo']}]"
                    role_yr_named = f"{role['advisory_named']} [{role['people_Advisory::YearFrom']} - {role['people_Advisory::YearTo']}]"
                    if role['people_Advisory::SCAR'] == 'Y':
                        scar.append(role_yr)
                        scar_named.append(role_yr_named)
                    else:
                        advisory.append(role_yr)
                        advisory_named.append(role_yr_named)

        if prsn.get('Supervises'):
            for reportee in prsn.get('Supervises'):
                if people.get(reportee):
                    position   = people.get(reportee)['Position']
                    salutation = people.get(reportee)['Salutation']
                    state      = people.get(reportee)['State'] 

                    if state == 'Active':
                        if position == 'Post Doc':
                            superPostdoc.append(salutation)
                        if position == 'Honours Student':
                            superHons.append(salutation)
                        if position == 'PhD Student':
                            superPhd.append(salutation)
                        if position == 'Masters Student':
                            superMasters.append(salutation)
                        if position == 'Associate Investigator':
                            superAssoc.append(salutation)
        
        context_idv['Prize']          = value_exists(prsn_prize)
        context_idv['Scar']           = value_exists(scar)  
        context_idv['Advisory']       = value_exists(advisory) 
        context_idv['ScarNamed']      = value_exists(scar_named)  
        context_idv['AdvisoryNamed']  = value_exists(advisory_named) 
        context_idv['PrizeNamed']     = value_exists(prize_named) 
        context_idv['SuperPostdoc']   = value_exists(superPostdoc)
        context_idv['SuperHons']      = value_exists(superHons)
        context_idv['SuperPhd']       = value_exists(superPhd)
        context_idv['SuperMasters']   = value_exists(superMasters)
        context_idv['SuperAssoc']     = value_exists(superAssoc)    

        prsn_res_outputs = res_outputs[(res_outputs.pubyr == str(yr)) & (res_outputs.id_person == id_prsn)]
        # This portion specifically addresses UOW's request to add project numbers to the bibiliography
        bucket_project = prsn_res_outputs[prsn_res_outputs.project != ''][['key', 'project', 'logistics']].drop_duplicates() 
        bucket_project.rename(columns={'key': 'ID_Zotero'}, inplace=True)
        biblio_project = bibliography.merge(bucket_project , how="left", on="ID_Zotero")
        biblio_project = biblio_project.fillna('na')
        biblio_project['Combined'] = biblio_project['Biblio'].where(biblio_project['project'] == 'na', other=biblio_project['Biblio'] \
            + '*** Project: '   + biblio_project['project'] + '  -  ' + 'Logistics: ' + biblio_project['logistics'] + '***')

        for x in prsn_res_outputs.iterrows():
            # create a bibliography per item type
            if x[1].itemType in prsn_output:
                prsn_output[x[1].itemType].append(get_biblio(biblio_project, x[1].key))
            else:
                prsn_output[x[1].itemType] = [get_biblio(biblio_project, x[1].key)] # add key in a list
            # create a bibliography per tag
            if bool([t for t in x[1].tags if t.get('tag') == 'to ngo']):
                ngo.append(get_biblio(biblio_project, x[1].key))
            if bool([t for t in x[1].tags if t.get('tag') == 'to industry']):
                industry.append(get_biblio(biblio_project, x[1].key))
            if bool([t for t in x[1].tags if t.get('tag') == 'to government']):
                govt.append(get_biblio(biblio_project, x[1].key))
            if bool([t for t in x[1].tags if t.get('tag') == 'to women in stem']):
                women.append(get_biblio(biblio_project, x[1].key))
            if bool([t for t in x[1].tags if t.get('tag') == 'to public']):
                public.append(get_biblio(biblio_project, x[1].key))
            if bool([t for t in x[1].tags if t.get('tag') == 'ats proposed papers']):
                ats.append(get_biblio(biblio_project, x[1].key))
            if bool([t for t in x[1].tags if t.get('tag') == 'museum engagement']):
                museum.append(get_biblio(biblio_project, x[1].key))
            if bool([t for t in x[1].tags if t.get('tag') == 'to professional bodies']):
                pro.append(get_biblio(biblio_project, x[1].key))

        # Outputs
        context_idv['Public']     = value_exists(public)
        context_idv['Women']      = value_exists(women)
        context_idv['Industry']   = value_exists(industry)
        context_idv['Ngo']        = value_exists(ngo)
        context_idv['Ats']        = value_exists(ats)
        context_idv['Govt']       = value_exists(govt)
        context_idv['Museum']     = value_exists(museum)
        context_idv['Pro']        = value_exists(pro)
        context_idv['Artwork']    = value_exists(prsn_output, 'artwork')
        context_idv['Plenary']    = value_exists(prsn_output, 'plenary')
        context_idv['Journal']    = value_exists(prsn_output, 'journalArticle')
        context_idv['Dataset']    = value_exists(prsn_output, 'dataset')
        context_idv['Book']       = "Book: None reported" if value_exists(prsn_output, 'book') is None else value_exists(prsn_output, 'book')
        context_idv['Chapter']    = "Chapter: None reported" if value_exists(prsn_output, 'bookSection') is None else value_exists(prsn_output, 'bookSection')
        context_idv['Conference'] = value_exists(prsn_output, 'conferencePaper')
        context_idv['Artwork']    = value_exists(prsn_output, 'artwork')
        context_idv['Film']       = value_exists(prsn_output, 'film')
        context_idv['Newspaper']  = value_exists(prsn_output, 'newspaperArticle')
        context_idv['Radio']      = value_exists(prsn_output, 'radioBroadcast')
        context_idv['Tv']         = value_exists(prsn_output, 'tvBroadcast')
        context_idv['Ntro']       = value_exists(prsn_output, 'ntro')
        context_idv['Report']     = unique_report(context_idv, value_exists(prsn_output, 'report'))
        context_idv['Present']    = unique_presentation(context_idv, value_exists(prsn_output, 'presentation'))
        # presentation only with zero impact on KPI counts
        context_idv['PrizePR']      = "None reported" if value_exists(prsn_prize) is None else value_exists(prsn_prize)
        context_idv['ScarPR']       = "None reported" if value_exists(scar) is None else value_exists(scar)
        context_idv['AdvisoryPR']   = "None reported" if value_exists(advisory) is None else value_exists(advisory)
        context_idv['PublicPR']     = "None reported" if value_exists(public) is None else value_exists(public)
        context_idv['WomenPR']      = "None reported" if value_exists(women) is None else value_exists(women)
        context_idv['IndustryPR']   = "None reported" if value_exists(industry) is None else value_exists(industry)
        context_idv['NgoPR']        = "None reported" if value_exists(ngo) is None else value_exists(ngo)
        context_idv['AtsPR']        = "None reported" if value_exists(ats) is None else value_exists(ats)
        context_idv['GovtPR']       = "None reported" if value_exists(govt) is None else value_exists(govt)
        context_idv['MuseumPR']     = "None reported" if value_exists(museum) is None else value_exists(museum)
        context_idv['ProPR']        = "None reported" if value_exists(pro) is None else value_exists(pro)
        context_idv['ArtworkPR']    = "None reported" if value_exists(prsn_output, 'artwork') is None else value_exists(prsn_output, 'artwork') 
        context_idv['PlenaryPR']    = "None reported" if value_exists(prsn_output, 'plenary') is None else value_exists(prsn_output, 'plenary') 
        context_idv['JournalPR']    = "None reported" if value_exists(prsn_output, 'journalArticle') is None else value_exists(prsn_output, 'journalArticle')
        context_idv['DatasetPR']    = "None reported" if value_exists(prsn_output, 'dataset') is None else value_exists(prsn_output, 'dataset')
        context_idv['BookPR']       = "Book: None reported" if value_exists(prsn_output, 'book') is None else value_exists(prsn_output, 'book')
        context_idv['ChapterPR']    = "Chapter: None reported" if value_exists(prsn_output, 'bookSection') is None else value_exists(prsn_output, 'bookSection')
        context_idv['ConferencePR'] = "None reported" if value_exists(prsn_output, 'conferencePaper') is None else value_exists(prsn_output, 'conferencePaper') 
        context_idv['ArtworkPR']    = "None reported" if value_exists(prsn_output, 'artwork') is None else value_exists(prsn_output, 'artwork') 
        context_idv['FilmPR']       = "None reported" if value_exists(prsn_output, 'film') is None else value_exists(prsn_output, 'film')
        context_idv['NewspaperPR']  = "None reported" if value_exists(prsn_output, 'newspaperArticle') is None else value_exists(prsn_output, 'newspaperArticle')
        context_idv['RadioPR']      = "None reported" if value_exists(prsn_output, 'radioBroadcast') is None else value_exists(prsn_output, 'radioBroadcast')
        context_idv['TvPR']         = "None reported" if value_exists(prsn_output, 'tvBroadcast') is None else value_exists(prsn_output, 'tvBroadcast') 
        context_idv['NtroPR']       = "None reported" if value_exists(prsn_output, 'ntro') is None else value_exists(prsn_output, 'ntro') 
        context_idv['ReportPR']     = "None reported" if unique_report(context_idv, value_exists(prsn_output, 'report')) is None else unique_report(context_idv, value_exists(prsn_output, 'report'))
        context_idv['PresentPR']    = "None reported" if unique_presentation(context_idv, value_exists(prsn_output, 'presentation')) is None else unique_presentation(context_idv, value_exists(prsn_output, 'presentation'))
 
    return(context_idv)

''' Create a dictionary of organisational content for an organisational MS Word template '''
def get_context_org(org, orgs, people, res_outputs, bibliography, yr, saef_projects):
    context_org_list                   = defaultdict(list)
    # context_org has strings, kpi_org has numbers
    context_org, kpi_org, pm_context   = {}, {}, {} 
    idv_list, org_summary, ppl_org,    = [], [], []
    program_members, ppl_projects      = [], []
    context_org['Organisation']  = orgs.get(org)
    kpi_org['kpiPublic']     = 0; kpi_org['kpiWomen']    = 0
    kpi_org['kpiGovernment'] = 0; kpi_org['kpiIndustry'] = 0
    kpi_org['kpiMuseum']     = 0; kpi_org['kpiPresent']  = 0
    kpi_org['kpiNgo']        = 0; kpi_org['kpiAts']      = 0
    kpi_org['Intern']        = 0; kpi_org['Volunteer']   = 0
    kpi_org['AssociateInv']  = 0; kpi_org['ResearchPro'] = 0
    kpi_org['ProgramStaff']  = 0; kpi_org['Position']    = 0
    kpi_org['ChiefOrPartnerInv'] = 0;
    HasProfile = ''
    
    pm_description = 'Description of work:\n\nPlease provide a \ndescription (~100 words minimum per SAEF\nproject) of the work the\n program member has been doing, listed by SAEF\n Project(s).'

    for prsn in people:
        p = people[prsn]
        if p['State'] == 'Active' and p['Org'] == org:
            if p['Position'] == 'Chief Investigator' or \
               p['Position'] == 'Partner Investigator':
                kpi_org['ChiefOrPartnerInv'] += 1
            if p['Position'] == 'Associate Investigator':
                kpi_org['AssociateInv']      += 1
            if p['Position'] == 'Research Professional':
                kpi_org['ResearchPro']       += 1
            if p['Position'] == 'Program Staff':
                kpi_org['ProgramStaff']      += 1
            if p['Position'] == 'Intern':
                kpi_org['Intern']            += 1
            if p['Position'] == 'Volunteer':
                kpi_org['Volunteer']         += 1

    for prsn in people:
        if people.get(prsn)['Org'] == org:
            idv = get_context_idv(org, people, prsn, res_outputs, bibliography, yr)
            idv_list.append(idv)
            ppl_org.append(prsn)

            if len(idv) > 0:
                program_members.append({'label': 'Organisation',               'cols': [orgs.get(org)], 'bg': '#D3D3D3'})
                program_members.append({'label': 'Position',                   'cols': [idv.get('Position')]})
                program_members.append({'label': 'Name',                       'cols': [idv.get('Salutation')]})
                program_members.append({'label': 'Start Date',                 'cols': [idv.get('StartDateDMY')]})
                program_members.append({'label': 'End Date',                   'cols': [idv.get('EndDateDMY')]})
                program_members.append({'label': 'SAEF FTE',                   'cols': [idv.get('Fte')]})
                program_members.append({'label': 'Project Number(s) with FTE', 'cols': [idv.get('ProjectCodeFTEList')]})
                program_members.append({'label': 'Profile',                    'cols': [idv.get('Profile')]})
                program_members.append({'label': 'Thesis',                     'cols': [idv.get('StudentProjectTitle')]})
                program_members.append({'label': 'Students: Confirmation of compliance with Participants Agreement Clause 25.3', 'cols': ['\n']})
                program_members.append({'label': 'Students: Cross-Node Supervision',           'cols': [idv.get('CrossnodeSupervision')]})
                program_members.append({'label': 'Students: Multidisciplinary Supervision',    'cols': ['\n']})
                program_members.append({'label': pm_description,               'cols': ['\n']})
                program_members.append({'label': 'ProfileURL',                 'cols': [idv.get('ProfileURL')]})
                program_members.append({'label': 'CareerStage',                'cols': [idv.get('CareerStage')]})
    
    pm_context['program_members'] = program_members

    # This portion specifically addresses UOW's request to add project numbers to the bibiliography
    bucket_project = res_outputs[res_outputs.project != ''][['key', 'project']].drop_duplicates() 
    bucket_project.rename(columns={'key': 'ID_Zotero'}, inplace=True)
    biblio_project = bibliography.merge(bucket_project , how="left", on="ID_Zotero")
    biblio_project = biblio_project.fillna('na')
    biblio_project['Combined'] = biblio_project['Biblio'].where(biblio_project['project'] == 'na', other=biblio_project['Biblio'] + ' *** Project: ' + biblio_project['project'] + ' ***')

    for i in idv_list:
        if len(i) > 0:
            if i['CareerStage'] == "Student":
                HasProfile = 'No' if len(i['StudentProjectTitle']) == 0 else 'Yes' 
            else:
                HasProfile = 'Yes' if i['ProfileURL'].startswith("https:") else 'No' # the presence of a url means we have a profile

            #     
            org_summary.append([i['Salutation'], i['Position'], i['StartDateDMY'], i['EndDateDMY'], \
                                i['Fte'], i['ProjectCodeFTEList'], i['lastname'], \
                                i['CareerStage'], i['CrossnodeSupervision'], HasProfile ])

            if i['ProjectCodeFTEList'] is not None:
                found = re.findall(r'T[0-9]+\_P[0-9]+', i['ProjectCodeFTEList'])
                if bool(found):
                    ppl_projects.append(found)

            context_org_list['Prize'].append(i['PrizeNamed'])
            if i['Scar'] is None:
                context_org_list['Scar'].append(i['Scar'])
            else:
                context_org_list['Scar'].append(i['ScarNamed'])
                
            if i['Advisory'] is None:
                context_org_list['Advisory'].append(i['Advisory'])
            else:
                context_org_list['Advisory'].append(i['AdvisoryNamed'])
                
            context_org_list['Ats'].append(i['Ats'])
            context_org_list['Workshop'].append(i['Workshop'])
            context_org_list['Ntro'].append(i['Ntro'])
            context_org_list['Public'].append(i['Public'])
            context_org_list['Women'].append(i['Women'])
            context_org_list['Govt'].append(i['Govt'])
            context_org_list['Industry'].append(i['Industry'])
            context_org_list['Ngo'].append(i['Ngo'])
            context_org_list['Plenary'].append(i['Plenary'])
            context_org_list['Museum'].append(i['Museum'])
            context_org_list['Pro'].append(i['Pro'])
            # Zotero tags
            kpi_org['kpiPublic']      += 0 if i['Public']   is None else 1
            kpi_org['kpiWomen']       += 0 if i['Women']    is None else 1
            kpi_org['kpiGovernment']  += 0 if i['Govt']     is None else 1
            kpi_org['kpiIndustry']    += 0 if i['Industry'] is None else 1
            kpi_org['kpiNgo']         += 0 if i['Ngo']      is None else 1
            kpi_org['kpiPresent']     += 0 if i['Present']  is None else 1
            kpi_org['kpiMuseum']      += 0 if i['Museum']   is None else 1
            kpi_org['kpiAts']         += 0 if i['Ats']      is None else 1
                        
        org_res_outputs =  res_outputs[res_outputs['id_person'].isin(ppl_org)]
        # create a datset that does not have any tags
        pr = org_res_outputs[(org_res_outputs.itemType == 'presentation') & (org_res_outputs.loc[:, ('tags')].apply(len)==0)]
       
        context_org['Journal']    = get_org_biblio(biblio_project, org_res_outputs, 'journalArticle')[0]
        context_org['Dataset']    = get_org_biblio(biblio_project, org_res_outputs, 'dataset')[0]
        context_org['Book']       = get_org_biblio(biblio_project, org_res_outputs, 'book')[0]
        context_org['Chapter']    = get_org_biblio(biblio_project, org_res_outputs, 'bookSection')[0]
        context_org['Conference'] = get_org_biblio(biblio_project, org_res_outputs, 'conferencePaper')[0]
        context_org['Report']     = get_org_biblio(biblio_project, org_res_outputs, 'report')[0]
        context_org['Artwork']    = get_org_biblio(biblio_project, org_res_outputs, 'artwork')[0]
        context_org['Film']       = get_org_biblio(biblio_project, org_res_outputs, 'film')[0]
        context_org['Newspaper']  = get_org_biblio(biblio_project, org_res_outputs, 'newspaperArticle')[0]
        context_org['Radio']      = get_org_biblio(biblio_project, org_res_outputs, 'radioBroadcast')[0]
        context_org['Tv']         = get_org_biblio(biblio_project, org_res_outputs, 'tvBroadcast')[0]
        context_org['Plenary']    = get_org_biblio(biblio_project, org_res_outputs, 'plenary')[0]
        context_org['Ntro']       = get_org_biblio(biblio_project, org_res_outputs, 'ntro')[0]
        context_org['Present']    = get_org_biblio(biblio_project, pr, 'presentation')[0]
        context_org['Prize']      = tidy_defaultdict(context_org_list['Prize'])
        context_org['Advisory']   = tidy_defaultdict(context_org_list['Advisory'])
        context_org['Scar']       = tidy_defaultdict(context_org_list['Scar'])
        context_org['Ats']        = tidy_defaultdict(context_org_list['Ats'])
        context_org['Workshop']   = tidy_defaultdict(context_org_list['Workshop'])
        context_org['Public']     = tidy_defaultdict(context_org_list['Public'],    'Public')
        context_org['Women']      = tidy_defaultdict(context_org_list['Women'],     'Women')
        context_org['Govt']       = tidy_defaultdict(context_org_list['Govt'],      'Govt')
        context_org['Industry']   = tidy_defaultdict(context_org_list['Industry'],  'Industry')
        context_org['Ngo']        = tidy_defaultdict(context_org_list['Ngo'],       'Ngo')
        context_org['Pro']        = tidy_defaultdict(context_org_list['Pro'],       'Pro')
        context_org['Museum']     = tidy_defaultdict(context_org_list['Museum'],    'Museum')

        # None makes more sense than Series([], ) if a Series is empty
        for x in ['Journal', 'Dataset', 'Book', 'Chapter', 'Conference', 'Report', 'Artwork', 'Film', 'Newspaper', 'Radio', 'Tv']:
            if context_org[x] == 'Series([], )':
                context_org[x] = 'None'

    kpi_org['kpiJournals']   = get_org_biblio(biblio_project, org_res_outputs, 'journalArticle')[1]
    kpi_org['kpiDataset']    = get_org_biblio(biblio_project, org_res_outputs, 'dataset')[1]
    kpi_org['kpiBook']       = get_org_biblio(biblio_project, org_res_outputs, 'book')[1]
    kpi_org['kpiTv']         = get_org_biblio(biblio_project, org_res_outputs, 'tvBroadcast')[1]
    kpi_org['kpiRadio']      = get_org_biblio(biblio_project, org_res_outputs, 'radioBroadcast')[1]
    kpi_org['kpiFilm']       = get_org_biblio(biblio_project, org_res_outputs, 'film')[1]
    kpi_org['kpiChapter']    = get_org_biblio(biblio_project, org_res_outputs, 'bookChapter')[1]
    kpi_org['kpiConference'] = get_org_biblio(biblio_project, org_res_outputs, 'conferencePaper')[1]
    kpi_org['kpiArtwork']    = get_org_biblio(biblio_project, org_res_outputs, 'artwork')[1]
    kpi_org['kpiNewspaper']  = get_org_biblio(biblio_project, org_res_outputs, 'newspaperArticle')[1]
    kpi_org['kpiReport']     = get_org_biblio(biblio_project, org_res_outputs, 'report')[1]
    kpi_org['kpiPresent']    = get_org_biblio(biblio_project, org_res_outputs, 'presentation')[1]
    kpi_org['kpiPlenary']    = get_org_biblio(biblio_project, org_res_outputs, 'plenary')[1]
    kpi_org['kpiNtro']       = get_org_biblio(biblio_project, org_res_outputs, 'ntro')[1]
    kpi_org['kpiPrize']      = value_count(context_org['Prize'])
    kpi_org['kpiScar']       = value_count(context_org['Scar'])
    kpi_org['kpiAdvisory']   = value_count(context_org['Advisory'])
    kpi_org['kpiPostDoc']    = sri_researcher_count(org_summary, 'Post Doc')
    kpi_org['kpiHonours']    = sri_researcher_count(org_summary, 'Honours Student')
    kpi_org['kpiPhd']        = sri_researcher_count(org_summary, 'PhD Student')
    kpi_org['kpiMasters']    = sri_researcher_count(org_summary, 'Masters Student')    

    # required data structure - # https://plumsail.com/docs/documents/v1.x/document-generation/docx/tables.html#regular-table
    projects = {'proj': []}
    
    for idx in saef_projects[saef_projects.ProjectLeadOrganisation == org].index:
    
        projects['proj'].append({'Code':   idx, \
                                 'Alias':  saef_projects['ProjectAlias'][idx], \
                                 'Title':  saef_projects['ProjectTitle'][idx], \
                                 'Status': saef_projects['Status'][idx], \
                                 'Contact':      saef_projects['Contact'][idx], \
                                 'Manager':      saef_projects['Manager'][idx], \
                                 'Organisation': saef_projects['ProjectLeadOrganisation'][idx] })

    projects_other = {'proj_oth': []}
    for idx in saef_projects[saef_projects.ProjectLeadOrganisation != org].index:
        if idx in list(chain.from_iterable(ppl_projects)):
            projects_other['proj_oth'].append({'Code':   idx, \
                                 'Alias':  saef_projects['ProjectAlias'][idx], \
                                 'Title':  saef_projects['ProjectTitle'][idx], \
                                 'Status': saef_projects['Status'][idx], \
                                 'Contact':      saef_projects['Contact'][idx], \
                                 'Organisation': saef_projects['ProjectLeadOrganisation'][idx] })
              
    context_org |= kpi_org
    context_org |= projects
    context_org |= projects_other
    context_org |= pm_context
    
    return  context_org, org_summary

''' Create a dictionary of program content for a program - SAEF,  MS Word template '''
def get_context_saef(organisations, people, res_outputs, bibliography, summary_bucket, yr, saef_projects):
    orgs = organisations.keys()
    saef = {'Honours': 0, 'Masters': 0, 'Phd':  0, 'PostDoc': 0, 'Prize': 0, 'Scar ': 0, \
            'ChiefOrPartnerInv': 0, 'AssociateInv': 0, 'ResearchPro': 0, 'ProgramStaff': 0, \
            'Intern': 0, 'Volunteer': 0 }
    
    zotero_tags = {'to public':        0, 'to industry':            0, 'museum engagement': 0,
                   'to women in stem': 0, 'to professional bodies': 0, 'to government':     0,
                   'to ngo':           0, 'ats proposed papers':    0}
    
    zotero_tags_nospace   = {}
    matched_key, saef_pm  = [], []
    value_counts = summary_bucket[summary_bucket.pubyr == str(yr)].itemType.value_counts().to_dict()
    bucket_tags  = res_outputs[(res_outputs.itemType == 'presentation') & (res_outputs.pubyr == str(yr))][['key', 'tags']]

    for zot_tag in zotero_tags:
        for tag in bucket_tags.iterrows(): 
            if len(tag[1][1]) > 0:
                if tag[1][1][0].get('tag') == zot_tag:
                    matched_key.append(tag[1][0]) # add matched key
        zotero_tags[zot_tag] = 0 if len(set(matched_key)) == 0 else len(set(matched_key))

    for tag in zotero_tags:
        zotero_tags_nospace[str.replace(tag,' ', '_')] = zotero_tags[tag]

    for o in orgs:
        co = get_context_org(o, organisations, people, res_outputs, bibliography, yr, saef_projects)[0]
        saef['PostDoc'] += co['kpiPostDoc']
        saef['Honours'] += co['kpiHonours']
        saef['Phd']     += co['kpiPhd']
        saef['Masters'] += co['kpiMasters']
        saef_pm.append(co['program_members'])

    # remove outer list
    saef['program_members'] = list(chain.from_iterable(saef_pm))

    for prsn in people:
        p = people[prsn]
        if p['State'] == 'Active':
            if p['Position'] == 'Chief Investigator' or p['Position'] == 'Partner Investigator':
                saef['ChiefOrPartnerInv'] += 1
            if p['Position'] == 'Associate Investigator':
                saef['AssociateInv'] += 1
            if p['Position'] == 'Research Professional':
                saef['ResearchPro'] += 1
            if p['Position'] == 'Program Staff':
                saef['ProgramStaff'] += 1
            if p['Position'] == 'Intern':
                saef['Intern'] += 1
            if p['Position'] == 'Volunteer':
                saef['Volunteer'] += 1
    projects = {'proj': []}
    for proj in saef_projects.iterrows():
        projects['proj'].append({'Code':  proj[0],     'Alias':        proj[1][1], \
                                 'Title':  proj[1][2], 'Organisation': proj[1][3], \
                                 'Status': proj[1][4], 'Name':         proj[1][0]})
            
    # concatenate dictionaries
    saef |= zotero_tags_nospace
    saef |= value_counts
    saef |= projects

    return saef

''' '''
# SAEF organisations
def write_context_org_excel(context_org, proj_saef, organisations, org, config_file='config/reporting.yaml'):
    with open(config_file, 'r') as file:
        cf = yaml.safe_load(file)

    if org not in ['Monash']:
        wb   = load_workbook(cf['templates']['org_template_excel'])
    elif org == 'Monash':
        wb   = load_workbook(cf['templates']['org_template_excel_xlarge'])
    ws1  = wb[cf['worksheets']['ws1']]
    ws4  = wb[cf['worksheets']['ws4']]

    context_org_idv = sorted(context_org[1], key=lambda x: x[6]) # Sort by lastName

    # Sheet: Organisation Summary

    # 1. Organisation Name
    ws1['D36'] = organisations.get(org)

    row_n=43

    # # 3. Personnel
    # Filter list by position
    omega = []
    a = [sublist for sublist in context_org_idv if sublist[1] in ("Chief Investigator", "Partner Investigator")]
    b = [sublist for sublist in context_org_idv if sublist[1] in ("Associate Investigator")]
    c = [sublist for sublist in context_org_idv if sublist[1] in ("Post Doc")]
    d = [sublist for sublist in context_org_idv if sublist[1] in ("PhD Student")]
    e = [sublist for sublist in context_org_idv if sublist[1] in ("Masters Student")]
    f = [sublist for sublist in context_org_idv if sublist[1] in ("Honours Student")]
    g = [sublist for sublist in context_org_idv if sublist[1] in ("Research Professional")]
    h = [sublist for sublist in context_org_idv if sublist[1] in ("Program Staff")]
    i = [sublist for sublist in context_org_idv if sublist[1] in ("Program Collaborator")]
    j = [sublist for sublist in context_org_idv if sublist[1] in ("Intern")]
    k = [sublist for sublist in context_org_idv if sublist[1] in ("Visitor")]

    if a:
        omega.append(a)
    if b:
        omega.append(b)
    if c:
        omega.append(c)
    if d:
        omega.append(d)
    if e:
        omega.append(e)
    if f:
        omega.append(f)
    if g:
        omega.append(g)
    if h:
        omega.append(h)
    if i:
        omega.append(i)
    if j:
        omega.append(j)
    if k:
        omega.append(k)

    for org_row in omega:
        for o in org_row:
            # ws1.cell(row=row_n, column=3, value=org_row[0])
            ws1['C'+str(row_n)] = o[0] # name including salutation [c43]
            ws1['D'+str(row_n)] = o[1] # saef position             [d43]
            ws1['E'+str(row_n)] = o[2] # start dt                  [e43]  
            ws1['F'+str(row_n)] = o[3] # end dt
            ws1['G'+str(row_n)] = o[4] # fte %
            ws1['H'+str(row_n)] = o[5] # project list
            ws1['I'+str(row_n)] = o[9] # has a profile/student thesis title
            row_n += 1

    if org not in ['Monash']:
        row_n = 91
    elif org == 'Monash':
        row_n = 108

    # # 4. Projects led by organisation
    context_org_proj = context_org[0]['proj']

    for p in context_org_proj:
        # 4. Projects led by the Organisation (4 rows)
        ws1['C'+str(row_n)] = p['Code']                         # ProjectCode
        ws1['D'+str(row_n)] = p['Alias']                        # ProjectAlias
        # ws1['C'+str(row_n)] = proj_org['ProjectTitle'][idx]   # ProjectTitle
        ws1.cell(row=row_n, column=5).value = p['Title']        # ProjectTitle
        ws1['H'+str(row_n)] = p['Contact']                      # Lead Investigator
        ws1.cell(row=row_n, column=9).value = p['Status']       # Project Approval Status?
        row_n += 1 

    # # Sheet: Key Performance Indicators

    # # 1. Number of research outputs
    ws4['D21'] = context_org[0]['kpiJournals']
    ws4['D22'] = context_org[0]['kpiDataset']
    ws4['D23'] = context_org[0]['kpiBook'] 
    ws4['D24'] = context_org[0]['kpiChapter'] 
    ws4['D25'] = context_org[0]['kpiConference']  
    # # ws4['D25'] = # ToDo: new Zotero tags needed

    # # 2. Quality of research outputs
    ws4['D33'] = context_org[0]['kpiPrize']
    # ws4['D34'] # ToDo: % of crossnode supervision 

    # # 5. Number of additional ARC-funded researchers working on SRI research
    ws4['D51'] = context_org[0]['kpiPostDoc']
    ws4['D52'] = context_org[0]['kpiHonours']
    ws4['D57'] = context_org[0]['kpiPhd']
    ws4['D58'] = context_org[0]['kpiMasters']

    # # 8. Number of presentations/briefings
    ws4['D67'] = context_org[0]['kpiPublic']
    ws4['D68'] = context_org[0]['kpiWomen']
    ws4['D69'] = context_org[0]['kpiGovernment']
    ws4['D70'] = context_org[0]['kpiIndustry']
    ws4['D71'] = context_org[0]['kpiNgo']
    # ws4['D69'] = context_org[0]['Pro']

    # # Centre-specific KPIs
    # ws4['D84'] = context_org[0]['Workshop']
    ws4['D89'] = context_org[0]['kpiArtwork']
    ws4['D90'] = context_org[0]['kpiAts'] # Antarctic Treaty System proposed papers
    ws4['D91'] = context_org[0]['kpiScar']
    ws4['D92'] = context_org[0]['kpiAdvisory']
    ws4['D95'] = context_org[0]['kpiFilm']
    ws4['D96'] = context_org[0]['kpiMuseum'] 
    ws4['D101'] = context_org[0]['kpiNewspaper'] 
    ws4['D102'] = context_org[0]['kpiRadio'] 
    ws4['D103'] = context_org[0]['kpiTv'] 

    wb.save(f"output/2025/midyear/org/{org}.xlsx")

''' Returns a surname: full name + project list dictionary
    Columns can be split useing a scolon as a delimiter '''
def get_project_list(saef_people):
    plist = {}
    for p in saef_people:
        if saef_people.get(p)['State'] == 'Active':
            name     = f"{saef_people.get(p)['FirstName']} {saef_people.get(p)['LastName']}"
            projects = sorted(saef_people.get(p)['Projects'])
            if len(projects) > 0:
                for proj in projects:
                    if saef_people.get(p)['LastName'] in plist:
                        plist[saef_people.get(p)['LastName']] = plist[saef_people.get(p)['LastName']] + ', ' + proj.split()[0].rstrip()
                    else:
                        plist[saef_people.get(p)['LastName']] = name + ':' + proj.split()[0].rstrip()
            else:
                plist[saef_people.get(p)['LastName']] = name

    # sort by surname
    plist_keys = list(plist.keys())
    plist_keys.sort()
    sorted_plist = {i: plist[i] for i in plist_keys}
    return sorted_plist

''' Return a list with values or a list with None (as a string)'''
def tidy_defaultdict(defaultdict_datatype, tag=None):
    # [i for i in context_org_list['Ats'] if i is not None]
    yy = [a for a in defaultdict_datatype if a is not None]
    if len(yy) > 0:
        return '\n\n'.join([''.join(e) for e in yy])
    else:
        return 'None reported'
  
''' Return a publication/author crosstab '''
def write_pub_auth_excel(res_outputs, saef_people, output="output/2024/org/pub_auth_crosstab.xlsx"):
    prsn_list = []
    for prsn in saef_people:
        if saef_people[prsn]['Position'] == 'Chief Investigator' or saef_people[prsn]['Position'] == 'Partner Investigator':
            prsn_list.append(prsn)

    yr  = get_rpt_args()[0]
    foo = res_outputs[res_outputs['id_person'].isin(prsn_list)] 
    pandas.crosstab(foo['title'], foo['name']).to_excel(output, sheet_name=f"{yr}")

def profile_exists(profile_name, url='https://arcsaef.com/researcher/', config_file='config/reporting.yaml'):
    
    with open(config_file, 'r') as file:
        cf = yaml.safe_load(file)

    saef_profile_exceptions  = json.loads(cf['exceptions']['saef_profile'])

    # remove diacritics or other language-specific symbols to their basic Latin equivalents
    if profile_name in saef_profile_exceptions.keys():
        profile_name_adjust = saef_profile_exceptions.get(profile_name)
        url = f"{url}{unidecode.unidecode(profile_name_adjust)}"
    else:
        url = f"{url}{unidecode.unidecode(profile_name)}"

    try:
        response = requests.get(url)
        if response.status_code == 200:
            return url
        else:
            return "Missing profile"
    except (requests.exceptions.HTTPError, requests.exceptions.ConnectionError) as e:
        return f"Failure - Unable to establish connection: {e}."
    except Exception as e:
        return f"Failure - Unknown error occurred: {e}."
